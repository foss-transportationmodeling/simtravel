'''
Created on Aug 17, 2011

@author: dhyou
'''

from PyQt4.QtGui import *
from PyQt4.QtCore import *
from openamos.gui.env import *
from openamos.core.database_management.cursor_database_connection import *
from openamos.core.database_management.database_configuration import *

import time
import numpy as np
import os
from copy import deepcopy

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_workbook
from openpyxl.drawing import Shape
from openpyxl.style import Color
from openpyxl.chart import LineChart, BarChart, Serie, ErrorBar, Reference

#from xlwt import *
#import win32com.client
#from win32com.client import Dispatch, constants

class Export_Outputs(QDialog):
    '''
    classdocs
    '''
    
    def __init__(self, configobject = None, parent = None):
        QDialog.__init__(self, parent)
        '''
        Constructor
        '''
        
        self.connects(configobject)
        self.setWindowTitle("Create basic OpenAmos outputs as MS Excel (.xls")
        self.setMinimumSize(QSize(350,300))
        alllayout = QVBoxLayout()
        self.setLayout(alllayout)
        
        
        excelinputbox = QGroupBox("")
        excellayout = QGridLayout()
        excelinputbox.setLayout(excellayout)
        
        selectwidget1 = QWidget()
        selectlayout1 = QVBoxLayout()
        selectwidget1.setLayout(selectlayout1)
        selectlayout1.setContentsMargins(0,0,0,40)
        
        csvnamelabel = QLabel("Save Outputs")
        selectlayout1.addWidget(csvnamelabel)
        
        filewidget = QWidget()
        filelayout = QHBoxLayout()
        filewidget.setLayout(filelayout)
        filelayout.setContentsMargins(0,0,0,0)
        self.xlsname = QLineEdit()
        self.xlsname.setMinimumWidth(200)
        filelayout.addWidget(self.xlsname)
        self.openfilebutton = QPushButton('...')
        self.openfilebutton.setMaximumWidth(30)
        filelayout.addWidget(self.openfilebutton)
        selectlayout1.addWidget(filewidget)
        
        tablenamelabel = QLabel("Select a person type")
        selectlayout1.addWidget(tablenamelabel)
        
        self.pptype = QComboBox()
        self.pptype.setMinimumWidth(200)
        self.pptype.addItems([QString("Adult Worker"),QString("Adult Non-worker"),QString("Non-adult (5-17)"),QString("Preschooler (0-4)")])
        selectlayout1.addWidget(self.pptype)
        
        
        selectwidget2 = QWidget()
        selectlayout2 = QVBoxLayout()
        selectwidget2.setLayout(selectlayout2)
        
        resultlabel = QLabel("Choose result(s)")
        self.resultchoice = QListWidget()
        self.resultchoice.setSelectionMode(QAbstractItemView.MultiSelection)
        #self.resultchoice.setFixedWidth(170)
        self.resultchoice.setMinimumHeight(150)
        vars = self.items()
        self.resultchoice.addItems(vars)
        selectlayout2.addWidget(resultlabel)
        selectlayout2.addWidget(self.resultchoice)
        

        #self.export_nhts = QCheckBox("Check if you would export NHTS instead of OpenAmos")
        self.isNHTS = False
        if (self.new_obj.check_if_table_exists("schedule_nhts") and self.new_obj.check_if_table_exists("trips_nhts") \
            and self.new_obj.check_if_table_exists("households_nhts") and self.new_obj.check_if_table_exists("persons_nhts") \
            and self.new_obj.check_if_table_exists("persons_daily_status_nhts")):
            
            self.isNHTS = True
        #    self.export_nhts.setDisabled(True)
            
        self.check_all = QCheckBox("Check to select all")

        
        excellayout.addWidget(selectwidget1,0,0)
        excellayout.addWidget(selectwidget2,0,1)
        #excellayout.addWidget(self.export_nhts,1,0)
        excellayout.addWidget(self.check_all,1,0)
    

        self.dialogButtonBox = QDialogButtonBox(QDialogButtonBox.Cancel| QDialogButtonBox.Ok)
        alllayout.addWidget(excelinputbox)
        alllayout.addWidget(self.dialogButtonBox)
        
        self.connect(self.openfilebutton, SIGNAL("clicked(bool)"), self.save_folder)
        self.connect(self.check_all, SIGNAL("stateChanged(int)"), self.select_all)
        self.connect(self.dialogButtonBox, SIGNAL("accepted()"), self, SLOT("accept()"))
        self.connect(self.dialogButtonBox, SIGNAL("rejected()"), self, SLOT("reject()"))

        
    def reject(self):
        self.disconnects()
        QDialog.accept(self)


    def accept(self):
        t1 = time.time()
        
        self.dialogButtonBox.setDisabled(True)
        filename = str(self.xlsname.text())
        sitems = self.resultchoice.selectedItems()
        if filename <> "" and len(sitems) > 0:
            
            
#        excel = win32com.client.Dispatch("Excel.Application")
#        book = excel.Workbooks.Add()
#        sheet = book.Worksheets(1)
#        sheet.Range("A1").Value = "Hello World!"
#        sheet.Range("A2").Value = str(Application.SIFilter(None, c.siObjectFilter))
#        book.SaveAs("c:\myBook.xls")
#         
#        sheet = None
#        book = None
#        excel.Quit()
#        excel = None 

                    #wsheet = wb.add_sheet(columns[i])  #xlwt code
#                    win32com code
#                    wsheet = wb.Worksheets.Add()
#                    wsheet.Name = str(columns[i])                        
#                    wb.Charts.Add()
#                    wb.ActiveChart.ChartType = 0x33
#                    wb.ActiveChart.SetSourceData(Source=wb.Sheets(str(columns[i])).Range(range("C3:C20"),range("G3:G20")))
#                    wb.ActiveChart.Location(Where = 0x2, Name=str(columns[i]))

#            win32com code
#            wsheet.Cells(i+3,j).Value = str(err1[i])
#            wsheet.Cells(i+3,j+1).Value = long(err2[i])
#            wsheet.Cells(i+3,j+2).Value = percent

#            xlwt code
#            wsheet.write(i+2,j,str(err1[i]))
#            wsheet.write(i+2,j+1,long(err2[i]))
#            wsheet.write(i+2,j+2,percent)

        
            wb = Workbook()
            
            columns = ["starttime","endtime","duration","trippurpose","starttime","endtime","duration","","dweltime","dweltime"]
            for i in range(self.resultchoice.count()):
                item = self.resultchoice.item(i)
                if item.isSelected() and i < 4:

                    wsheet = wb.create_sheet()
                    wsheet.title = columns[i]
                    
                    seri= []
                    seri.append(self.sql_quary1(wsheet,columns[i],False,i))
                    if self.isNHTS:
                        seri.append(self.sql_quary1(wsheet,columns[i],True,i))
                        self.call_chart(wsheet,seri)

                elif item.isSelected() and i >= 4 and i < 7:
                    
                    wsheet = wb.create_sheet()
                    wsheet.title = "%sbyPurpose"%(columns[i])
                                        
                    seri = []
                    seri.append(self.sql_quary2(wsheet,columns[i],False))
                    if self.isNHTS:
                        seri.append(self.sql_quary2(wsheet,columns[i],True))
                        self.call_chart(wsheet,seri)
                    
                elif item.isSelected() and (i == 7 or i == 11):

                    wsheet = wb.create_sheet()
                    if i == 7:
                        isTrip = True
                        wsheet.title = "TripRate"
                    else:
                        isTrip = False
                        wsheet.title = "ActivityRate"                    
                    
                    seri = []
                    seri.append(self.sql_quary3(wsheet,False,isTrip))
                    if self.isNHTS:
                        seri.append(self.sql_quary3(wsheet,True,isTrip))
                        self.call_chart(wsheet,seri)
                    
                elif item.isSelected() and i == 8:
                    
                    wsheet = wb.create_sheet()
                    wsheet.title = columns[i]
                    
                    seri = []
                    seri.append(self.sql_quary1(wsheet,columns[i],False,i))
                    if self.isNHTS:
                        seri.append(self.sql_quary1(wsheet,columns[i],True,i))
                        self.call_chart(wsheet,seri)
                    
                elif item.isSelected() and i == 9:
                    
                    wsheet = wb.create_sheet()
                    wsheet.title = "%sbyPurpose"%(columns[i])
                    
                    seri = []
                    seri.append(self.sql_quary2(wsheet,columns[i],False))
                    if self.isNHTS:
                        seri.append(self.sql_quary2(wsheet,columns[i],True))    
                        self.call_chart(wsheet,seri)
                    
                elif item.isSelected() and i == 10:
                    wsheet = wb.create_sheet()
                    wsheet.title = "TripRatebyPurpose"
                    seri = self.count_by_purpose(True,wsheet)
                    self.call_chart(wsheet,seri)
                
                elif item.isSelected() and i == 12:
                    wsheet = wb.create_sheet()
                    wsheet.title = "ActivityRatebyPurpose"
                    seri = self.count_by_purpose(False,wsheet)
                    self.call_chart(wsheet,seri)
                    
                elif item.isSelected() and i >= 13:
                    
                    wsheet = wb.create_sheet()
                    if i == 13:
                        column = "endtime"
                        wsheet.title = "EarliestStartDay"
                    else:
                        column = "starttime"
                        wsheet.title = "LatestEndDay"
                    
                    seri= []
                    seri.append(self.sql_quary1(wsheet,column,False,i))
                    if self.isNHTS:
                        seri.append(self.sql_quary1(wsheet,column,True,i))
                        self.call_chart(wsheet,seri)

            wb.save(os.path.realpath(filename))

            #QMessageBox.information(self, "",
            #    QString("""Outputs exporting is successful!"""), 
            #    QMessageBox.Yes)
            
            self.reject()
            
        else:
            QMessageBox.warning(self, "Save Outputs...",
                                QString("""Select at least one on the list."""), 
                                QMessageBox.Ok)            
        
        self.dialogButtonBox.setEnabled(True)
        t2 = time.time()
        print 'time taken --> %s'%(t2-t1)

                
    def save_folder(self):
        dialog = QFileDialog()
        filename = dialog.getSaveFileName(self,"Save File","","MS Excel 2007 (*.xlsx)")
        if filename <> "":
            if str(filename).rfind(".xlsx") < 0:
                filename = filename + ".xlsx"
            self.xlsname.setText(filename)

    def call_chart(self,wsheet,series):
        if self.isNHTS:
            temp = series[0]
            if len(temp) > 0:
                if type(temp[0]).__name__ == 'list':
                    seri1 = series[0]
                    seri2 = series[1]
                    chart_title = self.trip_labels("activitytype")
                    for i in range(len(seri1)):
                        if len(seri1[i]) == 4 or len(seri2[i]) == 4:
                            seri = [seri1[i],seri2[i]]
                            self.draw_chart(wsheet,seri,i*20,chart_title[i+1])
                else:
                    if len(series[0]) == 4 or len(series[1]) == 4:
                        self.draw_chart(wsheet,series)

    def draw_chart(self,wsheet,series,left=0,title=""):
        chart = BarChart()
        chart.title = title
        chart.drawing.left = 500+left
        chart.drawing.top = 250+left
        chart.drawing.height = 200
        chart.drawing.width = 500

        i = 0
        for serie in series:
            if len(serie) == 4:
                
                x1 = int(serie[0])
                y1 = int(serie[1])
                x2 = int(serie[2])
                y2 = int(serie[3])
                
                if x2 > 3:
                
                    if i == 0:
                        legend = Reference(wsheet,(0,0))
                        labels = Reference(wsheet,(x1,0),(x2,0))
                    else:
                        if y1 < 13:
                            legend = Reference(wsheet,(0,4))
                        if y1 >= 13:
                            legend = Reference(wsheet,(0,13))
                        
                    #value = wsheet.cell(row=4,column=0).value
                    title = wsheet.title
                    if title.find("Rate") < 0:
                        isLabel = True
                    else:
                        if len(title) > 12:
                            isLabel = True
                        else:
                            isLabel = False
                        
                    if i == 0 and isLabel: #type(value).__name__ <> 'int':
                        seri = Serie(Reference(wsheet,(x1,y1),(x2,y2)),labels=labels,legend=legend)
                    else:
                        seri = Serie(Reference(wsheet,(x1,y1),(x2,y2)),legend=legend)
                        
                    chart.add_serie(seri)
                
            i += 1

        wsheet.add_chart(chart)        


    def quary1(self,column,nhts):
        
        if nhts and column == "endtime":
            temp = "(select houseid, personid, min(%s) as %s from schedule_nhts where activitytype = 100 group by houseid, personid) as d"%(column,column)
        elif nhts and column == "starttime":
            temp = "(select houseid, personid, max(%s) as %s from schedule_nhts where activitytype = 100 group by houseid, personid) as d"%(column,column)
        elif not nhts and column == "endtime":
            temp = "(select houseid, personid, min(%s) as %s from schedule_full_r where activitytype = 100 group by houseid, personid) as d"%(column,column)
        else:
            temp = "(select houseid, personid, max(%s) as %s from schedule_full_r where activitytype = 100 group by houseid, personid) as d"%(column,column)
               
        return temp

    def per_quary(self,nhts):
        wrk = self.wrk_cond()
        if wrk == "0":
            wrk = "and (p.wrkdailystatus = 0 or p.wrkdailystatus is null)"
        elif wrk == "1":
            wrk = "and p.wrkdailystatus = 1"
        else:
            wrk = ""
        if not nhts:
            quary = "(select p.houseid, p.personid, p.age, p.wrkdailystatus from"
            quary = "%s (select ps.houseid, ps.personid, ps.age, w.wrkdailystatus from persons as ps"%(quary)
            quary = "%s left join persons_daily_status_r as w on ps.houseid = w.houseid and ps.personid = w.personid)"%(quary)
            quary = "%s as p where %s %s) as b"%(quary,self.age_cond(False),wrk)
        else:
            quary = "(select p.houseid, p.personid, p.r_age, p.wrkdailystatus, p.wtperfin from"
            quary = "%s (select ps.houseid, ps.personid, ps.r_age, ps.wtperfin, w.wrkdailystatus from persons_nhts as ps"%(quary)
            quary = "%s left join persons_daily_status_nhts as w on ps.houseid = w.houseid and ps.personid = w.personid)"%(quary)
            quary = "%s as p where %s %s) as b"%(quary,self.age_cond(True),wrk)

        return quary


    def sql_quary1(self,wsheet,column,nhts,cindex):
        
        nhts_var = ""
        per_wt = ""
        if nhts:# and column <> "dweltime":
            if column == "dweltime" or cindex >= 13:
                count = "sum(b.wtperfin)"
                per_wt = ", wtperfin"
            else:
                count = "sum(a.wttrdfin)"
                nhts_var = ", wttrdfin" 
        else:
            count = "count(*)"
            
        if column == "dweltime":
            column = "duration"
            tnames = self.tables(False)
        else:
            tnames = self.tables(True)
                
        if nhts:
            self.table_name(tnames)
            
        if cindex >= 13:
            tnames[0] = self.quary1(column,nhts)
        
#            for i in range(len(tnames)):
#                tnames[i] = tnames[i].replace("_r","") + "_nhts"

        
        cond = self.time_categroy(column)
        labels = self.trip_labels(column)
        temp = cond.keys()
        temp.sort()
        i = 1
        
        err1 = []
        err2 = []
        total = 0
        for key in temp:

            lowhigh = cond[key]
            sql = "select %s from "%(count)
            if len(lowhigh) > 1:
                sql = "%s(select houseid, personid%s from %s where %s >= %d and %s < %d order by houseid, personid) as a" %(sql,nhts_var,tnames[0],column,lowhigh[0],column,lowhigh[1])
            else:
                sql = "%s(select houseid, personid%s from %s where %s = %d) as a" %(sql,nhts_var,tnames[0],column,lowhigh[0])

            sql = "%s, %s where a.houseid = b.houseid and a.personid = b.personid"%(sql,self.per_quary(nhts))               
#            sql = "%s, (select houseid, personid%s from %s where %s order by houseid, personid) as b"%(sql,per_wt,tnames[1],self.age_cond(nhts))
#            wrk = self.wrk_cond()
#            if wrk != "":
#                if wrk == "0":
#                    wrk = "0 or wrkdailystatus is null"
#                sql = "%s, (select * from %s where wrkdailystatus = %s order by houseid, personid) as c"%(sql,tnames[2],wrk)
#                sql = "%s where a.houseid = b.houseid and a.personid = b.personid and a.houseid = c.houseid and a.personid = c.personid"%(sql)
#            else:
#                sql = "%s where a.houseid = b.houseid and a.personid = b.personid"%(sql)
            
            print sql
            self.new_obj.cursor.execute(sql)
            data = self.new_obj.cursor.fetchall()
            
            for j in data:
                i += 1
                if j[0] > 0:
                    err1.append(labels[key])
                    err2.append(long(j[0]))
                    total = total + long(j[0])
                else:
                    err1.append(labels[key])
                    err2.append(0)
        
        if nhts:
            j = 4
            self.cnames(wsheet,5)
        else:
            j = 0
            self.cnames(wsheet,1)
            
        for i in range(len(err1)):

            if total > 0.0:
                percent = round(100*float(err2[i])/total,2)
            else:
                percent = 0.0
            
            wsheet.cell(row=i+2,column=j).value = str(err1[i])
            wsheet.cell(row=i+2,column=j+1).value = long(err2[i])
            wsheet.cell(row=i+2,column=j+2).value = percent
        
        if total > 0:    
            location = [2,j+2,i+2,j+2]
        else:
            location = [-1]        
        return location

    def count_by_purpose(self,istrip,wsheet):

        if istrip:
            tnames = self.tables(True)
            purpose = "trippurpose"
        else:
            tnames = self.tables(False)
            purpose = "activitytype"
            
        
        cond = self.time_categroy(purpose)
        #ylabels = self.trip_labels(column)
        ykeys = cond.keys()
        ykeys.sort()
        
        yvalue = {}
        yvalue[0] = list(np.zeros(len(ykeys)))
        keyfreq = [0]
        cumulate = list(np.zeros(len(ykeys)))
        
        for key in ykeys:
            #wk = "(select * from %s order by houseid, personid) as d"%(tnames[2],self.wrk_cond())
            sql = "select a.freq, count(*) from %s"%(self.per_quary(False))
#            wk = self.wrk_cond()
#            if wk <> "":
#                if wk == "0":
#                    wk = "(d.wrkdailystatus = 0 or d.wrkdailystatus is null)"
#                else:
#                    wk = "d.wrkdailystatus = 1"
#                sql = "%s (select p.houseid, p.personid from %s as p, %s as d where p.houseid = d.houseid and p.personid = d.personid and %s and %s) as a"%(sql,tnames[1],tnames[2],self.age_cond(False),wk)
#            else:
#                sql = "%s (select p.houseid, p.personid from %s as p, %s as d where p.houseid = d.houseid and p.personid = d.personid and %s) as a"%(sql,tnames[1],tnames[2],self.age_cond(False))
                
            catecode = cond[key]
            if len(catecode) == 2:   
                where = "where %s >= %d and %s < %d"%(purpose,catecode[0],purpose,catecode[1])
            else:
                where = "where %s = %d"%(purpose,catecode[0])
                
            sql = "%s left join (select houseid, personid, count(*) as freq from %s %s group by houseid, personid) as a"%(sql,tnames[0],where)  
            sql = "%s on a.houseid = b.houseid and a.personid = b.personid group by a.freq order by a.freq"%(sql)
            print sql
             
            self.new_obj.cursor.execute(sql)
            rows = self.new_obj.cursor.fetchall()
            
            
            
            for row in rows:
                
                if row[0] <> None:
                    if int(row[0]) in keyfreq:
                        temp = yvalue[int(row[0])]
                        temp[key-1] = long(row[1])
                        cumulate[key-1] += long(row[1])
                    else:
                        temp = list(np.zeros(len(ykeys)))
                        temp[key-1] = long(row[1])
                        yvalue[int(row[0])] = temp
                        keyfreq.append(int(row[0]))
                        cumulate[key-1] += long(row[1])                       
                else:
                    temp = yvalue[0]
                    temp[key-1] = long(row[1])
                    cumulate[key-1] += long(row[1])
        
          
        print yvalue
        print cumulate
            

#        xlabels = self.trip_labels("trippurpose")
#        xkeys = xlabels.keys()
#        xkeys.sort()
#        
#        yvalue = {}
#        cumulate = np.zeros(len(xkeys))
#        cumulate = list(cumulate)
#        xvalue = np.zeros(len(xkeys))
#        xvalue = list(xvalue)
#        self.retrieve_twodimension(data, yvalue, cumulate, xkeys)


        if self.isNHTS:

            yvalue_nhts = {}
            yvalue_nhts[0] = list(np.zeros(len(ykeys)))
            keyfreq = [0]
            cumulate_nhts = list(np.zeros(len(ykeys)))
                    
            self.table_name(tnames)
            
            for key in ykeys:
                #wk = "(select * from %s order by houseid, personid) as d"%(tnames[2])
                sql = "select a.freq, sum(wtperfin) from %s"%(self.per_quary(True))
#                wk = self.wrk_cond()
#                if wk <> "":
#                    if wk == "0":
#                        wk = "(d.wrkdailystatus = 0 or d.wrkdailystatus is null)"
#                    else:
#                        wk = "d.wrkdailystatus = 1"
#                    sql = "%s (select p.houseid, p.personid, p.wtperfin from %s as p, %s as d where p.houseid = d.houseid and p.personid = d.personid and %s and %s) as a"%(sql,tnames[1],tnames[2],self.age_cond(True),wk)
#                else:
#                    sql = "%s (select p.houseid, p.personid, p.wtperfin from %s as p, %s as d where p.houseid = d.houseid and p.personid = d.personid and %s) as a"%(sql,tnames[1],tnames[2],self.age_cond(True))
                    
                catecode = cond[key]
                if len(catecode) == 2:   
                    where = "where %s >= %d and %s < %d"%(purpose,catecode[0],purpose,catecode[1])
                else:
                    where = "where %s = %d"%(purpose,catecode[0])
                
                sql = "%s left join (select houseid, personid, count(*) as freq from %s %s group by houseid, personid) as a"%(sql,tnames[0],where)  
                sql = "%s on a.houseid = b.houseid and a.personid = b.personid group by a.freq order by a.freq"%(sql)
                print sql
                 
                self.new_obj.cursor.execute(sql)
                rows = self.new_obj.cursor.fetchall()

                for row in rows:
                    
                    if row[0] <> None:
                        if int(row[0]) in keyfreq:
                            temp = yvalue_nhts[int(row[0])]
                            temp[key-1] = long(row[1])
                            cumulate_nhts[key-1] += long(row[1])
                        else:
                            temp = list(np.zeros(len(ykeys)))
                            temp[key-1] = long(row[1])
                            yvalue_nhts[int(row[0])] = temp
                            keyfreq.append(int(row[0]))
                            cumulate_nhts[key-1] += long(row[1])                       
                    else:
                        temp = yvalue_nhts[0]
                        temp[key-1] = long(row[1])
                        cumulate_nhts[key-1] += long(row[1])

            print yvalue_nhts
            print cumulate_nhts    


#            self.retrieve_twodimension(data, yvalue_nhts, cumulate_nhts, xkeys)
            
            key1 = yvalue.keys()
            key1.sort()
            key2 = yvalue_nhts.keys()
            key2.sort()
            
            xvalue = np.zeros(len(ykeys))
            xvalue = list(xvalue)
            
            for i in range(len(key1)):
                fre = key1[i]
                if fre not in key2:
                    yvalue_nhts[fre] = deepcopy(xvalue)
                    
            for i in range(len(key2)):
                fre = key2[i]
                if fre not in key1:
                    yvalue[fre] = deepcopy(xvalue)
        
        seri = []
        seri.append(self.outputs_twodimension(wsheet,False,yvalue,cumulate))
        if self.isNHTS:
            seri.append(self.outputs_twodimension(wsheet,True,yvalue_nhts,cumulate_nhts))
        
        return seri

    
#    def retrieve_twodimension(self,data,yvalue,cumulate,xkeys):    
#        xvalue = np.zeros(len(xkeys))
#        xvalue = list(xvalue)
#        
#        prate = 1
#        rate = None
#        for t in data:
#            if t[0] != None and t[1] != None:
#                
#                rate = int(t[0])
#                if prate <> rate:
#                    yvalue[prate] = deepcopy(xvalue)
#                    
#                    prate = rate
#                    xvalue = np.zeros(len(xkeys))
#                    xvalue = list(xvalue)
#                
#                temp = self.purpose_index(int(t[1]))
#                if temp > 0:
#                    index = xkeys.index(temp)
#                    xvalue[index] = xvalue[index] + long(t[2])
#                    cumulate[index] = cumulate[index] + long(t[2])
#            else:
#                ztrip = long(t[2])
#                ztrips =[]
#                for i in range(len(xkeys)): 
#                    ztrips.append(ztrip)
#                    cumulate[i] = cumulate[i] + ztrip
#                yvalue[0] = ztrips
#                
#        
#        if rate <> None: 
#            yvalue[rate] = deepcopy(xvalue)
        
        
    def outputs_twodimension(self,wsheet,nhts,yvalue,cumulate,column=""):

        xlabels = self.trip_labels("trippurpose")
        xkeys = xlabels.keys()
        xkeys.sort()
        
        if nhts:
            j = len(xkeys)+4
            wsheet.cell(row=0,column=len(xkeys)+3).value = "NHTS"            
        else:
            j = 1
            wsheet.cell(row=0,column=0).value = "OpenAmos"
            
        for xkey in xkeys:
            wsheet.cell(row=1,column=j).value = str(xlabels[xkey])
            j += 1
        
        if column <> "":
            ylabels = self.trip_labels(column)
            
        y = []
        keys = yvalue.keys()
        keys.sort()
        
        i = 2
        for index in range(len(keys)):
            
            if column <> "":
                ylabel = str(ylabels[keys[index]])
            else:
                ylabel = str(keys[index])
                
            if nhts:
                wsheet.cell(row=i,column=len(xkeys)+3).value = ylabel
                wsheet.cell(row=i+len(keys)+2,column=len(xkeys)+3).value = ylabel
            else:
                wsheet.cell(row=i,column=0).value = ylabel
                wsheet.cell(row=i+len(keys)+2,column=0).value = ylabel
                
            y = yvalue[keys[index]]
            for j in range(len(y)):
                if cumulate[j] > 0.0:
                    percent = round(100*float(y[j])/cumulate[j],2)
                else:
                    percent = 0.0

                if nhts:
                    wsheet.cell(row=i,column=j+len(xkeys)+4).value = percent
                    wsheet.cell(row=i+len(keys)+2,column=j+len(xkeys)+4).value = y[j]
                else:
                    wsheet.cell(row=i,column=j+1).value = percent
                    wsheet.cell(row=i+len(keys)+2,column=j+1).value = y[j]
                    
                
            i += 1
            
        locations = []
        for k in range(len(y)):
            if i-1 > 2:
                if nhts:
                    location = [i-len(yvalue),k+len(xkeys)+4,i-1,k+len(xkeys)+4]
                else:
                    location = [i-len(yvalue),k+1,i-1,k+1]
            else:
                location = []
                
            if cumulate[k] > 0.0:
                locations.append(location)
            else:
                locations.append([])
            
        return locations
    

    def sql_quary2(self,wsheet,column,nhts):

        nhts_var = ""
        per_wt = ""
        if nhts: # and column <> "dweltime":
            if column <> "dweltime":
                count = "sum(a.wttrdfin)"
                nhts_var = ", wttrdfin"
            else:
                count = "sum(b.wtperfin)"
                per_wt = ", wtperfin"                
        else:
            count = "count(*)"
            
        if column == "dweltime":
            column = "duration"
            tnames = self.tables(False)
            acttype = "activitytype"
        else:
            tnames = self.tables(True)
            acttype = "trippurpose"
        
            
        if nhts:
            self.table_name(tnames)
        
        
        cond = self.time_categroy(column)
        ylabels = self.trip_labels(column)
        ykeys = cond.keys()
        ykeys.sort()

        xlabels = self.trip_labels("trippurpose")
        xkeys = xlabels.keys()
        xkeys.sort()
        
        yvalue = {}
        cumulate = []
        for j in range(len(xkeys)):
            cumulate.append(0)
            
        for key in ykeys:
                
            lowhigh = cond[key]
            xvalue = []
            for j in range(len(xkeys)):
                xvalue.append(0)
                
            sql = "select %s, a.%s from "%(count,acttype)
            if len(lowhigh) > 1:
                sql = "%s(select houseid, personid, %s%s from %s where %s >= %d and %s < %d order by houseid, personid) as a" %(sql,acttype,nhts_var,tnames[0],column,lowhigh[0],column,lowhigh[1])
            else:
                sql = "%s(select houseid, personid, %s%s from %s where %s = %d) as a" %(sql,acttype,nhts_var,tnames[0],column,lowhigh[0])
            sql = "%s, %s where a.houseid = b.houseid and a.personid = b.personid"%(sql,self.per_quary(nhts))
#            sql = "%s, (select houseid, personid%s from %s where %s order by houseid, personid) as b"%(sql,per_wt,tnames[1],self.age_cond(nhts))
#            
#            wrk = self.wrk_cond()
#            if wrk != "":
#                if wrk == "0":
#                    wrk = "0 or wrkdailystatus is null"
#                sql = "%s, (select * from %s where wrkdailystatus = %s order by houseid, personid) as c"%(sql,tnames[2],wrk)
#                sql = "%s where a.houseid = b.houseid and a.personid = b.personid and a.houseid = c.houseid and a.personid = c.personid"%(sql)
#            else:
#                sql = "%s where a.houseid = b.houseid and a.personid = b.personid"%(sql)
            sql = "%s group by a.%s order by a.%s"%(sql,acttype,acttype)
            
            print sql
            self.new_obj.cursor.execute(sql)
            data = self.new_obj.cursor.fetchall()


            for t in data:
                if self.purpose_index(int(t[1])) > 0:
                    index = xkeys.index(self.purpose_index(int(t[1])))
                    xvalue[index] = xvalue[index] + long(t[0])
                    cumulate[index] = cumulate[index] + long(t[0])
            
            
            yvalue[key] = xvalue
            #yvalue[ylabels[key]] = xvalue
               
        return self.outputs_twodimension(wsheet,nhts,yvalue,cumulate,column)


    def sql_quary3(self,wsheet,nhts,istrip):
   
        tnames = self.tables(istrip)
        if nhts:
            self.table_name(tnames)
#            for i in range(len(tnames)):
#                tnames[i] = tnames[i].replace("_r","") + "_nhts"
                
#        wrk = self.wrk_cond()
#        if wrk == "":
#            table = "%s as p"%(tnames[1])
#            condition = ""
#        else:
#            if wrk == "0":
#                wrk = "(w.wrkdailystatus = 0 or w.wrkdailystatus is null)"
#            else:
#                wrk = "w.wrkdailystatus = 1"
#            table = "%s as p, %s as w"%(tnames[1],tnames[2])
#            condition = "p.houseid = w.houseid and p.personid = w.personid and %s and "%(wrk)
        
        nhts_var = ""
        if nhts:
            count = "sum(wtperfin)"
            nhts_var = ", p.wtperfin"
        else:
            count = "count(*)"
            
#        sql = "select a.freq, %s from (select p.houseid, p.personid%s from %s"%(count,nhts_var,table)
#        sql = "%s where %sp.%s) as b left join "%(sql,condition,self.age_cond(nhts))
        sql = "select a.freq, %s from %s left join"%(count,self.per_quary(nhts))
        sql = "%s (select houseid, personid, count(*) as freq from %s group by houseid, personid) as a "%(sql,tnames[0])
        sql = "%s on a.houseid = b.houseid and a.personid = b.personid group by a.freq order by a.freq"%(sql)

        
        print sql
        self.new_obj.cursor.execute(sql)
        data = self.new_obj.cursor.fetchall()
        
        err1 = []
        err2 = []
        total = 0
        for j in data:
            if j[0] != None:

                err1.append(int(j[0]))
                err2.append(long(j[1]))
                total += long(j[1])
            else:

                err1.insert(0,0)
                err2.insert(0,long(j[1]))
                total += long(j[1])

                
        
        if nhts:
            self.cnames(wsheet,5)
            j = 4 
        else:
            self.cnames(wsheet,1)
            j = 0
           
        i=0 
        k=0
        if len(err1) > 0:
            while i <= int(err1[len(err1)-1]):
                
                if i == err1[k]:
                    
                    if total > 0.0:
                        percent = round(100*float(err2[k])/total,2)
                    else:
                        percent = 0.0
                    
                    wsheet.cell(row=i+2,column=j).value = str(err1[k])
                    wsheet.cell(row=i+2,column=j+1).value = long(err2[k])
                    wsheet.cell(row=i+2,column=j+2).value = percent
                    k += 1
                else:
                    wsheet.cell(row=i+2,column=j).value = i
                    wsheet.cell(row=i+2,column=j+1).value = 0
                    wsheet.cell(row=i+2,column=j+2).value = 0
                
                i += 1
        
        if total > 0 and i+1 > 2:
            location = [2,j+2,i+1,j+2]
        else:
            location = [-1]
            
        return location               
                        
#        for i in range(len(err1)):
#            if total > 0.0:
#                percent = round(100*float(err2[i])/total,2)
#            else:
#                percent = 0.0
#                
#            wsheet.write(i+2,j,str(err1[i]))
#            wsheet.write(i+2,j+1,long(err2[i]))
#            wsheet.write(i+2,j+2,percent)

            


    def purpose_index(self, key):

        index = 0
        if key == 100:
            index = 1
        elif key >= 101 and key < 200:
            index = 2
        elif key >= 200 and key < 300:
            index = 3
        elif key >= 300 and key < 400:
            index = 4
        elif key >= 400 and key < 500:
            index = 5
        elif key >= 500 and key < 597:
            index = 6
        elif key >= 597 and key <=598:
            index = 7
        elif key == 600:
            index = 8
        elif key == 601:
            index = 9
        elif key == 599:
            index = 10
            
        return index                

    def age_cond(self,nhts):
        if self.pptype.currentIndex() <= 1:
            age = "p.age >= 18"
        elif self.pptype.currentIndex() == 2:
            age = "p.age < 18 and p.age >= 5"
        else:
            age = "p.age < 5"
            
        if nhts:
            age = age.replace("age", "r_age")
        return age

    def wrk_cond(self):
        if self.pptype.currentIndex() == 0:
            return "1"
        elif self.pptype.currentIndex() == 1:
            return "0"
        else:
            return ""
        
    def tables(self,istrip):
        if istrip:
            table_names = ["trips_purpose_r","persons","persons_daily_status_r"] #"daily_work_status_r"]
            return table_names
        else:
            table_names = ["schedule_full_r","persons","persons_daily_status_r"] #"daily_work_status_r"]
            return table_names 
    
    def table_name(self,names):
        for i in range(len(names)):
            if names[i].find("_full_r") > -1:
                names[i] = names[i].replace("_full_r","") + "_nhts"
            elif names[i].find("_purpose_r") > -1:
                names[i] = names[i].replace("_purpose_r","") + "_nhts"  
            elif names[i].find("_r") > -1:
                names[i] = names[i].replace("_r","") + "_nhts"  
            else:
                names[i] = names[i] + "_nhts"      

    def time_categroy(self,column):
        time = {1:[0,60],2:[60,120],3:[120,180],4:[180,240],5:[240,300],6:[300,360],
                7:[360,420],8:[420,480],9:[480,540],10:[540,600],11:[600,660],12:[660,720],
                13:[720,780],14:[780,840],15:[840,900],16:[900,960],17:[960,1020],18:[1020,1440]}
        mode = {1:[1],2:[2],3:[3],4:[4],5:[5],6:[6],7:[7],8:[8],9:[9],10:[10],11:[11]}
        duration = {1:[0,10],2:[10,20],3:[20,30],4:[30,50],5:[50,70],
                    6:[70,100],7:[100,150],8:[150,200],9:[200,250],10:[250,1440]}
        miles = {1:[0,5],2:[5,10],3:[10,15],4:[15,20],5:[20,25],
                 6:[25,30],7:[30,40],8:[40,50],9:[50,30000]}
        occupancy = {1:[0],2:[1],3:[2],4:[3],5:[4],6:[5,30000]}
        activitytype = {1:[100],2:[101,200],3:[200,300],4:[300,400],5:[400,500],
                        6:[500,597],7:[597,599],8:[600],9:[601],10:[599]}

        if column == "starttime" or column == "endtime": 
            return time
        elif column == "tripmode":
            return mode
        elif column == "duration":
            return duration
        elif column == "miles":
            return miles
        elif column == "occupancy":
            return occupancy
        elif column == "activitytype" or column == "trippurpose":
            return activitytype

    def trip_labels(self, column):
        activitytype = {1:'Home',2:'In-Home',3:'Work',4:'School',5:'Maintenance',6:'Discretionary',
                        7:'Anchor',8:'Pick Up',9:'Drop Off',10:'OH-Other'}             
        modedict = {1:'Car',2:'Van',3:'SUV',4:'Pickup Truck',5:'Bus',6:'Train',7:'School Bus',8:'Bike',9:'Walk',
                    10:'Taxi',11:'Other'}
        starttime = {1:'4am-5am',2:'5am-6am',3:'6am-7am',4:'7am-8am',5:'8am-9am',6:'9am-10am',
                    7:'10am-11am',8:'11am-12pm',9:'12pm-1pm',10:'1pm-2pm',11:'2pm-3pm',12:'3pm-4pm',
                    13:'4pm-5pm',14:'5pm-6pm',15:'6pm-7pm',16:'7pm-8pm',17:'8pm-9pm',18:'After 9pm'}
        endtime = {1:'4am-5am',2:'5am-6am',3:'6am-7am',4:'7am-8am',5:'8am-9am',6:'9am-10am',
                   7:'10am-11am',8:'11am-12pm',9:'12pm-1pm',10:'1pm-2pm',11:'2pm-3pm',12:'3pm-4pm',
                   13:'4pm-5pm',14:'5pm-6pm',15:'6pm-7pm',16:'7pm-8pm',17:'8pm-9pm',18:'After 9pm'}
        occupancy = {1:'0',2:'1',3:'2',4:'3',5:'4',6:'5 or more'}
        duration = {1:'0-10',2:'10-20',3:'20-30',4:'30-50',5:'50-70',
                    6:'70-100',7:'100-150',8:'150-200',9:'200-250',10:'>= 250'}
        miles = {1:'0-5',2:'5-10',3:'10-15',4:'15-20',5:'20-25',
                 6:'25-30',7:'30-40',8:'40-50',9:'>= 50'}
        
        if column == "activitytype" or column == "trippurpose":
            return activitytype
        if column == "tripmode":
            return modedict
        if column == "starttime":
            return starttime
        if column == "endtime":
            return endtime
        if column == "occupancy":
            return occupancy
        if column == "duration":
            return duration
        if column == "miles":
            return miles

    def select_all(self):
        if self.check_all.isChecked():
            for i in range(self.resultchoice.count()):
                temp = self.resultchoice.item(i)
                temp.setSelected(True)  
        else:
            for i in range(self.resultchoice.count()):
                temp = self.resultchoice.item(i)
                temp.setSelected(False)  
            
    def items(self):
        vars = ["trip starttime","trip endtime","trip duration","trip purpose","trip starttime by purpose",
                "trip endtime by purpose","trip duration by purpose","trip episode rate","dwell time","dwell time by activity",
                "trip rate by purpose","activity rate","activity rate by activity","earliest start of day","latest end of day"]
        return vars
    
    def tnames(self,wsheet,ind):
        names = self.items()
        wsheet.write(0,0,names[ind])

    def cnames(self,wsheet,i):
        if i == 1:
#            wsheet.write(0,i-1,"OpenAmos")
            wsheet.cell(row=0,column=i-1).value = "OpenAmos"
        else:
#            wsheet.write(0,i-1,"NHTS")
            wsheet.cell(row=0,column=i-1).value = "NHTS"
            
        wsheet.cell(row=1,column=i).value = "Frequency"
        wsheet.cell(row=1,column=i+1).value = "Percent(%)"
        
    def connects(self,configobject):
        
        protocol = configobject.getConfigElement(DB_CONFIG,DB_PROTOCOL)        
        user_name = configobject.getConfigElement(DB_CONFIG,DB_USER)
        password = configobject.getConfigElement(DB_CONFIG,DB_PASS)
        host_name = configobject.getConfigElement(DB_CONFIG,DB_HOST)
        database_name = configobject.getConfigElement(DB_CONFIG,DB_NAME)
        
        self.database_config_object = DataBaseConfiguration(protocol, user_name, password, host_name, database_name)
        self.new_obj = DataBaseConnection(self.database_config_object)
        self.new_obj.new_connection()
        

    def disconnects(self):
        self.new_obj.close_connection()


def main():
    app = QApplication(sys.argv)
    wizard = Export_Outputs()
    wizard.show()
    app.exec_()

if __name__=="__main__":
    main()
